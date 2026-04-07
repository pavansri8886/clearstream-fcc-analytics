"""
reports.py
----------
Reads alerts.csv directly and generates one Excel compliance report.
No intermediate files needed — run this right after pipeline.py.

Sheets:
  1. Summary   — headline KPI metrics
  2. Alerts    — top 1,000 HIGH severity alerts
  3. Sanctions — all sanctions hits

Usage:
    python src/reports.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "src"))

from config import CFG

ALERTS_PATH  = ROOT / CFG["paths"]["alerts_output"]
SUMMARY_PATH = ROOT / CFG["paths"]["summary_output"]
REPORT_PATH  = ROOT / CFG["paths"]["report_output"]

DARK_BLUE = "1F3864"
WHITE     = "FFFFFF"
GREY      = "F2F2F2"
RED       = "C00000"
AMBER     = "FF8C00"


def _load_data() -> tuple[pd.DataFrame, dict]:
    """Load alerts.csv and compute all KPIs in one place."""
    if not ALERTS_PATH.exists():
        print("alerts.csv not found. Run pipeline first: python src/pipeline.py")
        sys.exit(1)

    df = pd.read_csv(ALERTS_PATH, dtype=str, low_memory=False)
    df["amount_eur"] = pd.to_numeric(df["amount_eur"], errors="coerce").fillna(0)

    # Read total rows processed from pipeline summary
    rows_processed = 0
    if SUMMARY_PATH.exists():
        sp = pd.read_csv(SUMMARY_PATH)
        match = sp[sp["metric"] == "Total Rows Processed"]["value"]
        rows_processed = int(match.iloc[0]) if not match.empty else 0

    total = len(df)
    kpis = {
        "Total Rows Screened":       rows_processed,
        "Total Alerts":              total,
        "Alert Rate (%)":            round(total / max(rows_processed, 1) * 100, 2),
        "HIGH Severity":             int((df["alert_severity"] == "HIGH").sum()),
        "MEDIUM Severity":           int((df["alert_severity"] == "MEDIUM").sum()),
        "Sanctions Hits":            int((df["alert_type"] == "SANCTIONS_HIT").sum()),
        "Structuring Cases":         int((df["alert_type"] == "STRUCTURING").sum()),
        "Velocity Alerts":           int((df["alert_type"] == "VELOCITY_ABUSE").sum()),
        "Large Transaction Alerts":  int((df["alert_type"] == "LARGE_TRANSACTION").sum()),
        "High-Risk Corridor Alerts": int((df["alert_type"] == "HIGH_RISK_CORRIDOR").sum()),
        "SAR Candidates":            int((df["is_sar_candidate"].str.lower() == "true").sum()),
        "Total HIGH Exposure (EUR)": round(df[df["alert_severity"] == "HIGH"]["amount_eur"].sum(), 2),
    }

    return df, kpis


def generate() -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("pip install openpyxl")
        sys.exit(1)

    print("\nGenerating compliance report...")
    df, kpis = _load_data()
    print(f"  Loaded {len(df):,} alerts")

    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)

    # ── Shared style helpers ───────────────────────────────────────────────
    thin = Side(style="thin", color="BDD7EE")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_style(cell, bg=DARK_BLUE):
        cell.font      = Font(name="Calibri", bold=True, color=WHITE, size=11)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr

    def data_style(cell, alt=False):
        cell.font      = Font(name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = bdr
        if alt:
            cell.fill  = PatternFill("solid", fgColor=GREY)

    def write_headers(ws, headers: list, widths: list) -> None:
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            ws.column_dimensions[get_column_letter(col)].width = w
            cell = ws.cell(row=1, column=col, value=h)
            header_style(cell)
        ws.row_dimensions[1].height = 20

    def write_dataframe(ws, data: pd.DataFrame, start_row=2) -> None:
        for ri, (_, row) in enumerate(data.iterrows(), start_row):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                data_style(cell, alt=(ri % 2 == 0))
            ws.row_dimensions[ri].height = 15

    wb = Workbook()
    wb.remove(wb.active)

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1 — SUMMARY
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("Summary")
    ws1.sheet_view.showGridLines = False

    write_headers(ws1, ["Metric", "Value"], [36, 20])
    kpi_df = pd.DataFrame(list(kpis.items()), columns=["Metric", "Value"])
    write_dataframe(ws1, kpi_df)

    print(f"  Sheet 1: Summary ({len(kpis)} metrics)")

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 2 — TOP ALERTS (HIGH severity, sorted by amount)
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Alerts")
    ws2.freeze_panes = "A2"

    alert_cols   = ["alert_id", "alert_type", "alert_severity", "amount_eur",
                    "booking_date", "sender_name", "sender_country",
                    "receiver_name", "receiver_country", "description"]
    alert_widths = [18, 22, 14, 16, 14, 28, 14, 28, 14, 45]

    top = (
        df[df["alert_severity"] == "HIGH"]
        .sort_values("amount_eur", ascending=False)
        .head(1000)
        [alert_cols]
    )

    write_headers(ws2, alert_cols, alert_widths)
    write_dataframe(ws2, top)

    # Colour-code severity column (col 3)
    sev_colors = {"HIGH": RED, "MEDIUM": AMBER}
    for ri in range(2, len(top) + 2):
        sev = ws2.cell(row=ri, column=3).value
        if sev in sev_colors:
            ws2.cell(row=ri, column=3).font = Font(
                name="Calibri", bold=True, color=sev_colors[sev], size=10
            )

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(alert_cols))}1"
    print(f"  Sheet 2: Alerts ({len(top):,} HIGH severity alerts)")

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 3 — SANCTIONS HITS
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Sanctions")
    ws3.freeze_panes = "A2"

    san_cols   = ["alert_id", "transaction_id", "matched_value",
                  "matched_entity_name", "match_type", "match_score",
                  "list_source", "programme", "amount_eur",
                  "booking_date", "is_sar_candidate"]
    san_widths = [18, 22, 28, 28, 12, 10, 20, 18, 16, 14, 14]

    sanctions = df[df["alert_type"] == "SANCTIONS_HIT"][san_cols]

    write_headers(ws3, san_cols, san_widths)
    write_dataframe(ws3, sanctions)

    ws3.auto_filter.ref = f"A1:{get_column_letter(len(san_cols))}1"
    print(f"  Sheet 3: Sanctions ({len(sanctions):,} hits)")

    # ── Save ───────────────────────────────────────────────────────────────
    wb.save(REPORT_PATH)
    size_kb = REPORT_PATH.stat().st_size / 1024
    print(f"\nReport saved: {REPORT_PATH.name} ({size_kb:.0f} KB)")
    print(f"Sheets: {' | '.join(wb.sheetnames)}\n")


if __name__ == "__main__":
    generate()
