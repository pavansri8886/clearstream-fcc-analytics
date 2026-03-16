"""
report_generator.py
-------------------
Generates the FCC Compliance Report Excel workbook.
Reads from data/processed/summary/ — NOT from raw alerts.csv.
This separation means the report is fast and clean.

Output: reports/fcc_compliance_report.xlsx

Sheets:
  1. Executive Summary   — KPI cards + alert breakdown table
  2. Alert Log           — Top 1,000 alerts, filterable
  3. Sanctions Hits      — All sanctions hits with list source
  4. Monthly Trends      — Bar chart by alert type per month
  5. High-Risk Corridors — Top country pairs by exposure
  6. Regulatory Calendar — CSSF/FATF/EBA key dates 2025-2026

Usage:
    python src/report_generator.py
"""

from __future__ import annotations

import csv
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "src"))

SUMMARY_DIR  = ROOT / "data" / "processed" / "summary"
REPORT_PATH  = ROOT / "reports" / "fcc_compliance_report.xlsx"

# ── Colours ────────────────────────────────────────────────────────────────
C_DARK_BLUE  = "1F3864"
C_MID_BLUE   = "2E5DA6"
C_LIGHT_BLUE = "D6E4F0"
C_WHITE      = "FFFFFF"
C_RED        = "C00000"
C_AMBER      = "FF8C00"
C_GREEN      = "00B050"
C_GREY       = "F2F2F2"
C_BORDER     = "BDD7EE"

SEV_COLORS = {"HIGH": C_RED, "MEDIUM": C_AMBER, "LOW": C_GREEN}


def _load_csv(filename: str) -> list[dict]:
    path = SUMMARY_DIR / filename
    if not path.exists():
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def generate() -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        print("pip install openpyxl"); sys.exit(1)

    if not (SUMMARY_DIR / "kpis.csv").exists():
        print("Run aggregator first: python src/aggregator.py"); sys.exit(1)

    print("\n📊 Generating FCC Compliance Report...")
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)

    # Load all summary data
    kpis       = {r["metric"]: r["value"] for r in _load_csv("kpis.csv")}
    trends     = _load_csv("monthly_trends.csv")
    corridors  = _load_csv("country_exposure.csv")
    top_alerts = _load_csv("top_alerts.csv")
    sanctions  = [a for a in _load_csv("top_alerts.csv")
                  if a.get("alert_type") == "SANCTIONS_HIT"] + _load_csv("sar_candidates.csv")
    breakdown  = _load_csv("alert_type_breakdown.csv")

    wb = Workbook()
    wb.remove(wb.active)

    # ── Style helpers ──────────────────────────────────────────────────────
    def hfont(size=11, bold=True, color=C_WHITE):
        return Font(name="Calibri", bold=bold, color=color, size=size)
    def bfont(size=10, bold=False, color="000000"):
        return Font(name="Calibri", bold=bold, color=color, size=size)
    def fill(color):
        return PatternFill("solid", fgColor=color)
    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=False)
    thin = Side(style="thin", color=C_BORDER)
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_row(ws, row, n, color=C_DARK_BLUE):
        for c in range(1, n+1):
            cell = ws.cell(row=row, column=c)
            cell.font  = hfont()
            cell.fill  = fill(color)
            cell.alignment = center()
            cell.border = bdr

    def data_row(ws, row, n, alt=False):
        for c in range(1, n+1):
            cell = ws.cell(row=row, column=c)
            cell.font = bfont()
            cell.alignment = left()
            cell.border = bdr
            if alt:
                cell.fill = fill(C_GREY)

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1: EXECUTIVE SUMMARY
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("Executive Summary")
    ws1.sheet_view.showGridLines = False
    for col, w in [("A",4),("B",32),("C",18),("D",18),("E",18),("F",18)]:
        ws1.column_dimensions[col].width = w

    # Title
    ws1.merge_cells("B2:F2")
    tc = ws1["B2"]
    tc.value     = "CLEARSTREAM BANKING S.A. — FCC COMPLIANCE REPORT 2024"
    tc.font      = Font(name="Calibri", bold=True, color=C_WHITE, size=16)
    tc.fill      = fill(C_DARK_BLUE)
    tc.alignment = center()
    ws1.row_dimensions[2].height = 38

    ws1.merge_cells("B3:F3")
    sc = ws1["B3"]
    sc.value     = (f"Financial Crime Compliance | Generated: "
                    f"{datetime.today().strftime('%d %B %Y')} | "
                    f"Period: 01 Jan 2024 – 31 Dec 2024")
    sc.font      = Font(name="Calibri", italic=True, color=C_WHITE, size=10)
    sc.fill      = fill(C_MID_BLUE)
    sc.alignment = center()
    ws1.row_dimensions[3].height = 18

    # KPI cards — 2 per row
    kpi_items = [
        ("Transactions Screened",  kpis.get("Total Rows Screened","—"),  C_MID_BLUE, "MT103 + MT202 + MT540"),
        ("Total Alerts",           kpis.get("Total Alerts","—"),          C_RED,      "Require compliance review"),
        ("Sanctions Hits",         kpis.get("Sanctions Hits","—"),        C_RED,      "OFAC / UN / EU matches"),
        ("SAR Candidates",         kpis.get("SAR Candidates","—"),        C_RED,      "Filed to goAML / CSSF"),
        ("Structuring Cases",      kpis.get("Structuring Cases","—"),     "7030A0",   "Sub-threshold splitting"),
        ("Alert Rate",             f"{kpis.get('Alert Rate (%)','—')}%",  C_AMBER,    "% of transactions flagged"),
    ]

    row = 5
    col_pairs = [(2, 4), (2, 4), (2, 4)]
    for i, (label, value, color, note) in enumerate(kpi_items):
        col = 2 if i % 2 == 0 else 4
        if i > 0 and i % 2 == 0:
            row += 5

        # Label bar
        ws1.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(col+1)}{row}")
        c = ws1.cell(row=row, column=col, value=label.upper())
        c.font = hfont(size=10); c.fill = fill(color); c.alignment = center(); c.border = bdr
        ws1.row_dimensions[row].height = 18

        # Value
        ws1.merge_cells(f"{get_column_letter(col)}{row+1}:{get_column_letter(col+1)}{row+1}")
        c = ws1.cell(row=row+1, column=col, value=value)
        c.font = Font(name="Calibri", bold=True, color=color, size=24)
        c.fill = fill(C_LIGHT_BLUE); c.alignment = center(); c.border = bdr
        ws1.row_dimensions[row+1].height = 38

        # Note
        ws1.merge_cells(f"{get_column_letter(col)}{row+2}:{get_column_letter(col+1)}{row+2}")
        c = ws1.cell(row=row+2, column=col, value=note)
        c.font = Font(name="Calibri", italic=True, color="666666", size=9)
        c.alignment = center()
        ws1.row_dimensions[row+2].height = 14

    # Alert type breakdown table
    tbl_row = row + 6
    ws1.merge_cells(f"B{tbl_row}:F{tbl_row}")
    c = ws1.cell(row=tbl_row, column=2, value="ALERT BREAKDOWN BY TYPE")
    c.font = hfont(size=12); c.fill = fill(C_DARK_BLUE); c.alignment = center()
    ws1.row_dimensions[tbl_row].height = 22

    hdrs = ["Alert Type", "Count", "% of Total", "Severity", "FATF Typology"]
    for ci, h in enumerate(hdrs, 2):
        ws1.cell(row=tbl_row+1, column=ci, value=h)
    header_row(ws1, tbl_row+1, len(hdrs)+1, C_MID_BLUE)

    for i, r in enumerate(breakdown):
        ri = tbl_row + 2 + i
        vals = [r.get("alert_type",""), r.get("count",""),
                f"{r.get('pct_of_total',0)}%", r.get("severity",""),
                r.get("fatf_typology","")]
        for ci, v in enumerate(vals, 2):
            cell = ws1.cell(row=ri, column=ci, value=v)
            if ci == 5:   # severity
                sev_raw = str(v).split("–")[0].strip()
                cell.font = Font(name="Calibri", bold=True,
                                 color=SEV_COLORS.get(sev_raw, "000000"), size=10)
            data_row(ws1, ri, len(vals)+1, alt=(i%2==1))
        ws1.row_dimensions[ri].height = 16

    print("  ✓ Sheet 1: Executive Summary")

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 2: ALERT LOG
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Alert Log")
    ws2.freeze_panes = "A2"

    a_cols = [
        ("Alert ID",18),("Transaction ID",22),("Type",22),
        ("Severity",12),("Amount EUR",16),("Date",14),
        ("Sender",28),("Sender Cty",12),
        ("Receiver",28),("Receiver Cty",12),
        ("Corridor",16),("Status",12),("Description",45),
    ]
    for ci, (h, w) in enumerate(a_cols, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.cell(row=1, column=ci, value=h)
    header_row(ws2, 1, len(a_cols))
    ws2.row_dimensions[1].height = 20

    a_fields = [
        "alert_id","transaction_id","alert_type","alert_severity",
        "amount_eur","booking_date","sender_name","sender_country",
        "receiver_name","receiver_country","corridor","alert_status","description",
    ]
    for ri, row_d in enumerate(top_alerts[:1000], 2):
        for ci, f in enumerate(a_fields, 1):
            v = row_d.get(f, "")
            if f == "amount_eur":
                try: v = round(float(v), 2)
                except: v = 0
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.font = bfont(); cell.alignment = left(); cell.border = bdr
            if f == "alert_severity":
                cell.font = Font(name="Calibri", bold=True,
                                 color=SEV_COLORS.get(str(v),"000000"), size=10)
        if ri % 2 == 0:
            for ci in range(1, len(a_cols)+1):
                ws2.cell(row=ri, column=ci).fill = fill(C_GREY)
        ws2.row_dimensions[ri].height = 14

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(a_cols))}1"
    print("  ✓ Sheet 2: Alert Log")

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 3: SANCTIONS HITS
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Sanctions Hits")
    ws3.freeze_panes = "A2"

    s_cols = [
        ("Alert ID",18),("Transaction ID",22),("Matched Name",30),
        ("Sanctioned Entity",30),("Match Type",14),("Score",10),
        ("List Source",20),("Programme",18),("Country",12),
        ("Amount EUR",16),("Date",14),("SAR?",8),
    ]
    for ci, (h, w) in enumerate(s_cols, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
        ws3.cell(row=1, column=ci, value=h)
    header_row(ws3, 1, len(s_cols), C_RED)
    ws3.row_dimensions[1].height = 20

    seen_san = set()
    unique_sanctions = []
    for r in sanctions:
        tid = r.get("transaction_id","")
        if tid not in seen_san:
            seen_san.add(tid)
            unique_sanctions.append(r)

    s_fields = [
        "alert_id","transaction_id","matched_value","matched_entity_name",
        "match_type","match_score","list_source","programme","sanctions_country",
        "amount_eur","booking_date","is_sar_candidate",
    ]
    for ri, row_d in enumerate(unique_sanctions[:500], 2):
        for ci, f in enumerate(s_fields, 1):
            v = row_d.get(f, "")
            if f == "amount_eur":
                try: v = round(float(v), 2)
                except: v = 0
            cell = ws3.cell(row=ri, column=ci, value=v)
            cell.font = bfont(); cell.alignment = left(); cell.border = bdr
        if ri % 2 == 0:
            for ci in range(1, len(s_cols)+1):
                ws3.cell(row=ri, column=ci).fill = fill(C_GREY)
        ws3.row_dimensions[ri].height = 14

    ws3.auto_filter.ref = f"A1:{get_column_letter(len(s_cols))}1"
    print("  ✓ Sheet 3: Sanctions Hits")

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 4: MONTHLY TRENDS
    # ══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Monthly Trends")
    if trends:
        trend_cols = list(trends[0].keys())
        for ci, h in enumerate(trend_cols, 1):
            ws4.column_dimensions[get_column_letter(ci)].width = 18
            ws4.cell(row=1, column=ci, value=h)
        header_row(ws4, 1, len(trend_cols), C_DARK_BLUE)

        for ri, row_d in enumerate(trends, 2):
            for ci, col in enumerate(trend_cols, 1):
                v = row_d.get(col, 0)
                try:
                    v = int(v) if col != "month" else v
                except: pass
                cell = ws4.cell(row=ri, column=ci, value=v)
                cell.font = bfont(); cell.alignment = left(); cell.border = bdr
            data_row(ws4, ri, len(trend_cols), alt=(ri%2==0))
            ws4.row_dimensions[ri].height = 16

        # Bar chart
        if len(trends) > 1 and len(trend_cols) > 1:
            chart = BarChart()
            chart.type = "col"; chart.title = "Monthly Alert Volume by Type"
            chart.y_axis.title = "Count"; chart.style = 10
            chart.width = 26; chart.height = 14
            data_ref = Reference(ws4, min_col=2, max_col=len(trend_cols),
                                 min_row=1, max_row=len(trends)+1)
            cats    = Reference(ws4, min_col=1, min_row=2, max_row=len(trends)+1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            ws4.add_chart(chart, f"A{len(trends)+5}")

    print("  ✓ Sheet 4: Monthly Trends")

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 5: HIGH RISK CORRIDORS
    # ══════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("High-Risk Corridors")
    c5_cols = [("Corridor",22),("Alert Count",16),("Total Exposure EUR",22)]
    for ci, (h, w) in enumerate(c5_cols, 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w
        ws5.cell(row=1, column=ci, value=h)
    header_row(ws5, 1, len(c5_cols), "833C00")

    for ri, row_d in enumerate(corridors[:30], 2):
        vals = [row_d.get("corridor",""),
                int(float(row_d.get("alert_count",0) or 0)),
                round(float(row_d.get("total_exposure_eur",0) or 0), 2)]
        for ci, v in enumerate(vals, 1):
            cell = ws5.cell(row=ri, column=ci, value=v)
            cell.font = bfont(); cell.alignment = left(); cell.border = bdr
        data_row(ws5, ri, 3, alt=(ri%2==0))
        ws5.row_dimensions[ri].height = 16

    print("  ✓ Sheet 5: High-Risk Corridors")

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 6: REGULATORY CALENDAR
    # ══════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Regulatory Calendar")
    reg_cols = [
        ("Date",14),("Regulation",28),("Regulator",18),
        ("Obligation",48),("Status",14),("Owner",22),
    ]
    for ci, (h, w) in enumerate(reg_cols, 1):
        ws6.column_dimensions[get_column_letter(ci)].width = w
        ws6.cell(row=1, column=ci, value=h)
    header_row(ws6, 1, len(reg_cols), C_DARK_BLUE)

    calendar = [
        ("Jan 2025",  "DORA (Art.17-23)",        "EBA/CSSF",      "ICT incident reporting framework in force",             "COMPLETE",     "CRO / IT Risk"),
        ("Mar 2025",  "EU AMLR",                  "EBA",           "Enhanced CDD for high-risk jurisdictions effective",    "IN PROGRESS",  "Head of AML"),
        ("Jun 2025",  "FATF R.16 (updated)",      "FATF/CSSF",     "Travel Rule — cross-border payment data requirements",  "IN PROGRESS",  "Compliance PMO"),
        ("Jul 2025",  "AMLA Establishment",        "AMLA",          "EU AML Authority begins operations in Frankfurt",       "MONITORING",   "Head of Compliance"),
        ("Sep 2025",  "CSSF Questionnaire",        "CSSF",          "Annual AML/CFT questionnaire submission deadline",      "COMPLETE",     "Compliance PMO"),
        ("Oct 2025",  "AMLD6 Transposition",       "Luxembourg MoF","6th AML Directive national law deadline",              "IN PROGRESS",  "Legal"),
        ("Dec 2025",  "DORA Full Application",     "EBA/EIOPA",     "ICT third-party risk management — full compliance",     "NOT STARTED",  "CRO"),
        ("Jan 2026",  "AMLD6 In Force",            "CSSF",          "Beneficial ownership registers updated",                "NOT STARTED",  "Head of AML"),
        ("Mar 2026",  "CSSF Risk Assessment",      "CSSF",          "Biennial AML/CFT risk assessment submission",           "NOT STARTED",  "Compliance PMO"),
        ("Jul 2026",  "AMLA Direct Supervision",   "AMLA",          "AMLA begins direct supervision of selected entities",   "MONITORING",   "Group CCO"),
    ]
    status_colors_map = {
        "COMPLETE":    C_GREEN, "IN PROGRESS": C_AMBER,
        "NOT STARTED": C_RED,   "MONITORING":  C_MID_BLUE,
    }
    for ri, row_data in enumerate(calendar, 2):
        for ci, v in enumerate(row_data, 1):
            cell = ws6.cell(row=ri, column=ci, value=v)
            cell.font = bfont(); cell.alignment = left(); cell.border = bdr
            if ci == 5:
                cell.font = Font(name="Calibri", bold=True,
                                 color=status_colors_map.get(str(v), "000000"), size=10)
        data_row(ws6, ri, len(reg_cols), alt=(ri%2==0))
        ws6.row_dimensions[ri].height = 20

    print("  ✓ Sheet 6: Regulatory Calendar")

    # Save
    wb.save(REPORT_PATH)
    size = REPORT_PATH.stat().st_size / 1024
    print(f"\n✅ Report saved: {REPORT_PATH.name} ({size:.0f} KB)")
    print(f"   Sheets: {' | '.join(wb.sheetnames)}\n")


if __name__ == "__main__":
    generate()
