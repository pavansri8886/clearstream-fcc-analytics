"""
sar_generator.py
----------------
Generates the SAR (Suspicious Activity Report) Register Excel workbook.
Reads from data/processed/summary/sar_candidates.csv.

The SAR Register is a core compliance document — it tracks every potential
STR (Suspicious Transaction Report) filed or under consideration.
At Clearstream, STRs are filed to Luxembourg's FIU (CRF) via goAML.

This document is what a compliance analyst would hand to the
Head of AML when sanctions hits require escalation.

Output: reports/sar_register.xlsx

Sheets:
  1. SAR Register     — All SAR candidates with narrative and status
  2. Filing Checklist — CSSF/goAML submission requirements per SAR
  3. Regulatory Notes — Luxembourg STR filing obligations

Usage:
    python src/sar_generator.py
"""

from __future__ import annotations

import csv
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "src"))

SAR_PATH    = ROOT / "data" / "processed" / "summary" / "sar_candidates.csv"
REPORT_PATH = ROOT / "reports" / "sar_register.xlsx"

C_DARK_BLUE = "1F3864"
C_MID_BLUE  = "2E5DA6"
C_LIGHT_BLUE= "D6E4F0"
C_WHITE     = "FFFFFF"
C_RED       = "C00000"
C_AMBER     = "FF8C00"
C_GREEN     = "00B050"
C_GREY      = "F2F2F2"
C_BORDER    = "BDD7EE"


def _load_sar_candidates() -> list[dict]:
    if not SAR_PATH.exists():
        print(f"SAR candidates not found: {SAR_PATH}")
        print("Run: python src/aggregator.py")
        sys.exit(1)
    with open(SAR_PATH, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def generate() -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("pip install openpyxl"); sys.exit(1)

    candidates = _load_sar_candidates()
    print(f"\n🚨 Generating SAR Register — {len(candidates)} candidates...")
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Styles ─────────────────────────────────────────────────────────────
    def hfont(size=11, bold=True, color=C_WHITE):
        return Font(name="Calibri", bold=bold, color=color, size=size)
    def bfont(size=10, bold=False, color="000000"):
        return Font(name="Calibri", bold=bold, color=color, size=size)
    def fill(color):
        return PatternFill("solid", fgColor=color)
    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=C_BORDER)
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_header(ws, row, n, color=C_DARK_BLUE):
        for c in range(1, n+1):
            cell = ws.cell(row=row, column=c)
            cell.font = hfont(); cell.fill = fill(color)
            cell.alignment = center(); cell.border = bdr

    def set_data(ws, row, n, alt=False):
        for c in range(1, n+1):
            cell = ws.cell(row=row, column=c)
            cell.font = bfont(); cell.alignment = left(); cell.border = bdr
            if alt:
                cell.fill = fill(C_GREY)

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1: SAR REGISTER
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("SAR Register")
    ws1.freeze_panes = "A2"
    ws1.sheet_view.showGridLines = False

    col_widths = {
        "A": 16, "B": 22, "C": 28, "D": 28,
        "E": 14, "F": 14, "G": 20, "H": 16,
        "I": 16, "J": 20, "K": 18, "L": 18,
    }
    for col, w in col_widths.items():
        ws1.column_dimensions[col].width = w

    # Title
    ws1.merge_cells("A1:L1")
    tc = ws1["A1"]
    tc.value = (f"CLEARSTREAM BANKING S.A. — SUSPICIOUS ACTIVITY REPORT REGISTER  |  "
                f"Generated: {datetime.today().strftime('%d %B %Y')}  |  "
                f"CONFIDENTIAL — RESTRICTED")
    tc.font  = Font(name="Calibri", bold=True, color=C_WHITE, size=12)
    tc.fill  = fill(C_RED)
    tc.alignment = center()
    ws1.row_dimensions[1].height = 24

    # Headers
    hdrs = [
        "SAR Reference", "Transaction ID", "Matched Name",
        "Sanctioned Entity", "List Source", "Programme",
        "Amount EUR", "Transaction Date", "Filing Status",
        "Filed To", "Filed Date", "Analyst Note",
    ]
    for ci, h in enumerate(hdrs, 1):
        ws1.cell(row=2, column=ci, value=h)
    set_header(ws1, 2, len(hdrs), C_RED)
    ws1.row_dimensions[2].height = 20

    # SAR reference counter
    sar_num = 1
    for ri, row_d in enumerate(candidates[:200], 3):
        sar_ref   = f"SAR-2024-{sar_num:04d}"
        amount    = row_d.get("amount_eur","")
        try:
            amount = round(float(amount), 2)
        except: pass

        # Determine filing status based on match type and amount
        match_type = row_d.get("match_type","")
        try:
            amt_val = float(row_d.get("amount_eur",0) or 0)
        except: amt_val = 0

        if match_type == "EXACT" and amt_val > 100_000:
            filing_status = "FILED"
            filed_to      = "CSSF / CRF (goAML)"
            filed_date    = row_d.get("booking_date","")
        elif match_type == "EXACT":
            filing_status = "UNDER REVIEW"
            filed_to      = "Pending CCO approval"
            filed_date    = ""
        else:
            filing_status = "MONITORING"
            filed_to      = "Internal — no filing yet"
            filed_date    = ""

        vals = [
            sar_ref,
            row_d.get("transaction_id",""),
            row_d.get("matched_value",""),
            row_d.get("matched_entity_name",""),
            row_d.get("list_source",""),
            row_d.get("programme",""),
            amount,
            row_d.get("booking_date",""),
            filing_status,
            filed_to,
            filed_date,
            f"Auto-flagged by FCC screening system. {row_d.get('match_type','')} match score: {row_d.get('match_score','')}",
        ]

        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci, value=v)
            cell.font = bfont(); cell.alignment = left(); cell.border = bdr
            if ci == 9:   # Filing status
                status_color = {
                    "FILED":        C_RED,
                    "UNDER REVIEW": C_AMBER,
                    "MONITORING":   C_MID_BLUE,
                }.get(str(v), "000000")
                cell.font = Font(name="Calibri", bold=True, color=status_color, size=10)

        set_data(ws1, ri, len(hdrs), alt=(ri%2==0))
        ws1.row_dimensions[ri].height = 14
        sar_num += 1

    ws1.auto_filter.ref = f"A2:{get_column_letter(len(hdrs))}2"
    print("  ✓ Sheet 1: SAR Register")

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 2: FILING CHECKLIST
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Filing Checklist")
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 30

    ws2.merge_cells("B2:D2")
    tc = ws2["B2"]
    tc.value = "STR FILING CHECKLIST — Luxembourg goAML Requirements"
    tc.font  = hfont(size=13); tc.fill = fill(C_DARK_BLUE)
    tc.alignment = center()
    ws2.row_dimensions[2].height = 26

    ws2.merge_cells("B3:D3")
    sc = ws2["B3"]
    sc.value = "Pursuant to: Luxembourg Law of 12 November 2004 | CSSF Regulation 12-02 | EU AMLR"
    sc.font  = Font(name="Calibri", italic=True, color=C_WHITE, size=10)
    sc.fill  = fill(C_MID_BLUE); sc.alignment = center()
    ws2.row_dimensions[3].height = 16

    checklist_hdrs = ["Requirement", "Mandatory?", "Notes"]
    for ci, h in enumerate(checklist_hdrs, 2):
        ws2.cell(row=5, column=ci, value=h)
    set_header(ws2, 5, 3, C_MID_BLUE)

    checklist_items = [
        ("Suspicious transaction identified by compliance system", "MANDATORY", "Automated alert from FCC pipeline"),
        ("Initial review by AML analyst completed", "MANDATORY", "Within 24 hours of alert generation"),
        ("Transaction blocked / funds frozen if required", "MANDATORY", "For confirmed OFAC/UN hits — immediate"),
        ("Compliance Officer notified", "MANDATORY", "CCO must approve STR before filing"),
        ("STR drafted using goAML template", "MANDATORY", "Narrative must include: who, what, when, why suspicious"),
        ("Supporting documentation attached", "MANDATORY", "Transaction records, KYC documents, screening evidence"),
        ("STR filed to CRF (Luxembourg FIU) via goAML", "MANDATORY", "Within 30 days of suspicion arising"),
        ("Simultaneous notification to CSSF", "MANDATORY", "Same day as CRF filing for sanctions-related STRs"),
        ("OFAC notification (for OFAC-programme hits)", "IF APPLICABLE", "Required for US sanctions — OFAC reporting portal"),
        ("Internal SAR Register updated", "MANDATORY", "This document — track filing date and reference"),
        ("Client tipping-off prohibition observed", "MANDATORY", "Do NOT inform client that STR has been filed"),
        ("Record kept for minimum 5 years", "MANDATORY", "EU AML Directive — extended to 7 years for Clearstream"),
        ("Quarterly SAR register review with CCO", "GOOD PRACTICE", "Review pending and monitoring cases"),
    ]

    for ri, (req, mandatory, notes) in enumerate(checklist_items, 6):
        ws2.cell(row=ri, column=2, value=req)
        ws2.cell(row=ri, column=3, value=mandatory)
        ws2.cell(row=ri, column=4, value=notes)

        color = C_RED if mandatory == "MANDATORY" else C_AMBER
        ws2.cell(row=ri, column=3).font = Font(name="Calibri", bold=True, color=color, size=10)

        set_data(ws2, ri, 3, alt=(ri%2==0))
        ws2.row_dimensions[ri].height = 18

    print("  ✓ Sheet 2: Filing Checklist")

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 3: REGULATORY NOTES
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Regulatory Notes")
    ws3.column_dimensions["A"].width = 5
    ws3.column_dimensions["B"].width = 60

    ws3.merge_cells("B2:B2")
    tc = ws3["B2"]
    tc.value = "REGULATORY FRAMEWORK — STR OBLIGATIONS AT CLEARSTREAM BANKING S.A."
    tc.font  = hfont(size=13, color=C_WHITE); tc.fill = fill(C_DARK_BLUE)
    tc.alignment = left()
    ws3.row_dimensions[2].height = 26

    notes = [
        ("LEGAL BASIS",
         "Luxembourg Law of 12 November 2004 on AML/CFT (as amended) requires "
         "Clearstream Banking S.A., as a regulated financial institution, to file a "
         "Suspicious Transaction Report (STR) — referred to as a Déclaration d'Opération "
         "Suspecte (DOS) in French — with the Cellule de Renseignement Financier (CRF), "
         "Luxembourg's Financial Intelligence Unit, whenever there is a suspicion or "
         "reasonable grounds to suspect money laundering or terrorist financing."),
        ("TIMING REQUIREMENT",
         "STRs must be filed as soon as possible after the suspicion arises. For sanctions "
         "hits involving UN Security Council designations, the expectation is action within "
         "hours. Simultaneous notification to the CSSF (Commission de Surveillance du Secteur "
         "Financier) is required for sanctions-related reports."),
        ("TIPPING-OFF PROHIBITION",
         "Article 5 of the 2004 Law strictly prohibits informing the subject of an STR that "
         "a report has been filed or that an investigation is underway. Breach of the tipping-off "
         "prohibition is a criminal offence under Luxembourg law."),
        ("GOAML PLATFORM",
         "All STRs are filed electronically via the goAML platform operated by the CRF. "
         "Clearstream's compliance team has direct access. Each filing receives a unique "
         "CRF reference number which must be recorded in this SAR Register."),
        ("DATA RETENTION",
         "STR records, supporting documentation, and the SAR Register must be retained for "
         "a minimum of 5 years under EU AML Directive requirements. Clearstream's internal "
         "policy extends this to 7 years to align with broader record-keeping obligations."),
        ("EU AML PACKAGE (2025-2026)",
         "Under the new EU AML Regulation (AMLR) and AMLD6, STR obligations are being "
         "harmonised across the EU. The new EU Anti-Money Laundering Authority (AMLA) will "
         "begin direct supervision of Clearstream as a selected obliged entity from July 2026. "
         "Enhanced STR reporting standards under AMLA supervision are expected."),
    ]

    row = 4
    for heading, text in notes:
        # Section heading
        hc = ws3.cell(row=row, column=2, value=heading)
        hc.font  = Font(name="Calibri", bold=True, color=C_WHITE, size=11)
        hc.fill  = fill(C_MID_BLUE); hc.alignment = left(); hc.border = bdr
        ws3.row_dimensions[row].height = 18
        row += 1

        # Text
        tc = ws3.cell(row=row, column=2, value=text)
        tc.font      = Font(name="Calibri", size=10, color="222222")
        tc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        tc.border    = bdr
        ws3.row_dimensions[row].height = 72
        row += 2

    print("  ✓ Sheet 3: Regulatory Notes")

    # Save
    wb.save(REPORT_PATH)
    size = REPORT_PATH.stat().st_size / 1024
    print(f"\n✅ SAR Register saved: {REPORT_PATH.name} ({size:.0f} KB)")
    print(f"   Sheets: {' | '.join(wb.sheetnames)}\n")


if __name__ == "__main__":
    generate()
