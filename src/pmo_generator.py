"""
pmo_generator.py
----------------
Generates the FCC Programme PMO Tracker Excel workbook.
Reads project data from config/pmo_projects.yaml.

This is the Track 3 deliverable — the document a compliance PMO intern
would actually produce and hand to management on a Monday morning.

Output: reports/fcc_project_status.xlsx

Sheets:
  1. Programme Dashboard  — RAG overview + progress bars
  2. Project Details      — Milestone tracking per project
  3. Action Log           — Open actions and blockers
  4. Regulatory Deadlines — Linked from pmo_projects.yaml

Usage:
    python src/pmo_generator.py
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "src"))

PMO_CONFIG   = ROOT / "config" / "pmo_projects.yaml"
REPORT_PATH  = ROOT / "reports" / "fcc_project_status.xlsx"

C_DARK_BLUE  = "1F3864"
C_MID_BLUE   = "2E5DA6"
C_LIGHT_BLUE = "D6E4F0"
C_WHITE      = "FFFFFF"
C_RED        = "C00000"
C_AMBER      = "FF8C00"
C_GREEN      = "00B050"
C_GREY       = "F2F2F2"
C_BORDER     = "BDD7EE"

RAG_COLORS   = {"RED": C_RED, "AMBER": C_AMBER, "GREEN": C_GREEN}
STATUS_COLORS = {
    "COMPLETE":     C_GREEN,
    "IN_PROGRESS":  C_AMBER,
    "NOT_STARTED":  C_RED,
    "ON_HOLD":      C_MID_BLUE,
}


def _load_projects() -> list[dict]:
    try:
        import yaml
    except ImportError:
        print("pip install pyyaml"); sys.exit(1)

    if not PMO_CONFIG.exists():
        print(f"PMO config not found: {PMO_CONFIG}"); sys.exit(1)

    with open(PMO_CONFIG, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    return data.get("projects", [])


def generate() -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        print("pip install openpyxl"); sys.exit(1)

    projects = _load_projects()
    print(f"\n📋 Generating PMO Tracker — {len(projects)} projects...")
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Style helpers ──────────────────────────────────────────────────────
    def hfont(size=11, bold=True, color=C_WHITE):
        return Font(name="Calibri", bold=bold, color=color, size=size)
    def bfont(size=10, bold=False, color="000000"):
        return Font(name="Calibri", bold=bold, color=color, size=size)
    def fill(color):
        return PatternFill("solid", fgColor=color)
    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=C_BORDER)
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_header(ws, row, n, color=C_DARK_BLUE):
        for c in range(1, n+1):
            cell = ws.cell(row=row, column=c)
            cell.font = hfont(); cell.fill = fill(color)
            cell.alignment = center(); cell.border = bdr

    def set_data(ws, row, n, alt=False):
        for c in range(1, n+1):
            cell = ws.cell(row=row, column=c)
            cell.font = bfont(); cell.alignment = left(); cell.border = bdr
            if alt:
                cell.fill = fill(C_GREY)

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1: PROGRAMME DASHBOARD
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("Programme Dashboard")
    ws1.sheet_view.showGridLines = False

    col_widths = {"A":5,"B":12,"C":32,"D":24,"E":20,"F":10,"G":16,"H":20,"I":38}
    for col, w in col_widths.items():
        ws1.column_dimensions[col].width = w

    # Title
    ws1.merge_cells("B2:I2")
    tc = ws1["B2"]
    tc.value     = "FCC PROGRAMME STATUS DASHBOARD — CLEARSTREAM BANKING S.A."
    tc.font      = Font(name="Calibri", bold=True, color=C_WHITE, size=15)
    tc.fill      = fill(C_DARK_BLUE)
    tc.alignment = center()
    ws1.row_dimensions[2].height = 34

    ws1.merge_cells("B3:I3")
    sc = ws1["B3"]
    sc.value     = (f"Reporting Date: {datetime.today().strftime('%d %B %Y')}  |  "
                    f"Programme: Financial Crime Compliance  |  "
                    f"Prepared by: FCC PMO Team")
    sc.font      = Font(name="Calibri", italic=True, color=C_WHITE, size=10)
    sc.fill      = fill(C_MID_BLUE)
    sc.alignment = center()
    ws1.row_dimensions[3].height = 18

    # RAG summary row
    total_p  = len(projects)
    red_p    = sum(1 for p in projects if p.get("rag_status") == "RED")
    amber_p  = sum(1 for p in projects if p.get("rag_status") == "AMBER")
    green_p  = sum(1 for p in projects if p.get("rag_status") == "GREEN")
    complete = sum(1 for p in projects if p.get("status") == "COMPLETE")

    summary_boxes = [
        ("TOTAL PROJECTS", total_p,  C_MID_BLUE),
        ("RED",            red_p,    C_RED),
        ("AMBER",          amber_p,  C_AMBER),
        ("GREEN",          green_p,  C_GREEN),
        ("COMPLETE",       complete, C_GREEN),
    ]
    col = 2
    for label, val, color in summary_boxes:
        ws1.merge_cells(f"{get_column_letter(col)}5:{get_column_letter(col)}5")
        lc = ws1.cell(row=5, column=col, value=label)
        lc.font = hfont(size=9); lc.fill = fill(color); lc.alignment = center(); lc.border = bdr
        ws1.row_dimensions[5].height = 16

        vc = ws1.cell(row=6, column=col, value=val)
        vc.font = Font(name="Calibri", bold=True, color=color, size=20)
        vc.fill = fill(C_LIGHT_BLUE); vc.alignment = center(); vc.border = bdr
        ws1.row_dimensions[6].height = 30
        col += 1

    # Project table header
    tbl_row = 9
    hdrs = ["Project ID", "Project Name", "Owner", "Regulation",
            "% Complete", "RAG", "End Date", "Blockers"]
    for ci, h in enumerate(hdrs, 2):
        ws1.cell(row=tbl_row, column=ci, value=h)
    set_header(ws1, tbl_row, len(hdrs)+1, C_MID_BLUE)
    ws1.row_dimensions[tbl_row].height = 20

    for i, p in enumerate(projects):
        ri = tbl_row + 1 + i
        pct    = p.get("pct_complete", 0)
        rag    = p.get("rag_status", "AMBER")
        status = p.get("status", "IN_PROGRESS")
        row_data = [
            p.get("id",""), p.get("name",""), p.get("owner",""),
            p.get("regulation","")[:30], f"{pct}%", rag,
            p.get("end_date",""), p.get("blockers","")[:50],
        ]
        for ci, v in enumerate(row_data, 2):
            cell = ws1.cell(row=ri, column=ci, value=v)
            cell.font   = bfont()
            cell.alignment = left()
            cell.border = bdr
            if ci == 7:   # RAG column
                cell.font = Font(name="Calibri", bold=True,
                                 color=RAG_COLORS.get(str(v), "000000"), size=12)
                cell.alignment = center()
        set_data(ws1, ri, len(row_data)+1, alt=(i%2==1))
        ws1.row_dimensions[ri].height = 20

    print("  ✓ Sheet 1: Programme Dashboard")

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 2: PROJECT DETAILS + MILESTONES
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Project Details")
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 42
    ws2.column_dimensions["D"].width = 24
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 14
    ws2.column_dimensions["G"].width = 12

    row = 2
    for p in projects:
        # Project header block
        ws2.merge_cells(f"B{row}:G{row}")
        c = ws2.cell(row=row, column=2,
                     value=f"{p.get('id','')} — {p.get('name','')}")
        c.font = hfont(size=12); c.fill = fill(C_DARK_BLUE)
        c.alignment = left(); c.border = bdr
        ws2.row_dimensions[row].height = 22
        row += 1

        # Meta row
        meta = [
            ("Owner", p.get("owner","")),
            ("Regulation", p.get("regulation","")),
            ("Regulator", p.get("regulator","")),
            ("Status", p.get("status","").replace("_"," ")),
            ("RAG", p.get("rag_status","")),
            ("% Complete", f"{p.get('pct_complete',0)}%"),
        ]
        for ci, (label, val) in enumerate(meta, 2):
            lc = ws2.cell(row=row, column=ci, value=label)
            lc.font = hfont(size=9, color=C_WHITE)
            lc.fill = fill(C_MID_BLUE); lc.alignment = center(); lc.border = bdr
            vc = ws2.cell(row=row+1, column=ci, value=val)
            vc.font = bfont(bold=True); vc.alignment = center(); vc.border = bdr
            if label == "RAG":
                vc.font = Font(name="Calibri", bold=True,
                               color=RAG_COLORS.get(str(val),"000000"), size=11)
        ws2.row_dimensions[row].height = 16
        ws2.row_dimensions[row+1].height = 18
        row += 2

        # Description
        ws2.merge_cells(f"B{row}:G{row}")
        dc = ws2.cell(row=row, column=2, value=str(p.get("description","")).strip())
        dc.font = Font(name="Calibri", italic=True, size=9, color="444444")
        dc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws2.row_dimensions[row].height = 28
        row += 1

        # Milestones header
        ms_hdrs = ["#", "Milestone", "Due Date", "Done"]
        for ci, h in enumerate(ms_hdrs, 2):
            c = ws2.cell(row=row, column=ci, value=h)
            c.font = hfont(size=9); c.fill = fill(C_MID_BLUE)
            c.alignment = center(); c.border = bdr
        ws2.row_dimensions[row].height = 16
        row += 1

        for mi, ms in enumerate(p.get("milestones", []), 1):
            done = ms.get("done", False)
            vals = [mi, ms.get("name",""), ms.get("due",""),
                    "✓ Done" if done else "○ Pending"]
            for ci, v in enumerate(vals, 2):
                cell = ws2.cell(row=row, column=ci, value=v)
                cell.font  = bfont(
                    color=C_GREEN if done else C_AMBER,
                    bold=(ci==5)
                )
                cell.alignment = left(); cell.border = bdr
            ws2.row_dimensions[row].height = 16
            row += 1

        # Blockers row
        bc = ws2.cell(row=row, column=2, value=f"⚠ Blockers: {p.get('blockers','None')}")
        bc.font = Font(name="Calibri", bold=True,
                       color=C_RED if p.get("blockers","None") != "None" else C_GREEN,
                       size=9)
        bc.border = bdr
        ws2.row_dimensions[row].height = 16
        row += 2   # gap between projects

    print("  ✓ Sheet 2: Project Details")

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 3: ACTION LOG
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Action Log")
    ws3.freeze_panes = "A2"

    act_cols = [
        ("Action ID",14),("Project",16),("Action",45),
        ("Owner",22),("Due Date",14),("Priority",12),("Status",14),
    ]
    for ci, (h, w) in enumerate(act_cols, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
        ws3.cell(row=1, column=ci, value=h)
    set_header(ws3, 1, len(act_cols), C_DARK_BLUE)

    # Build action items from project blockers and upcoming milestones
    actions = []
    action_num = 1
    for p in projects:
        pid = p.get("id","")

        # Add blocker as action if not "None"
        if p.get("blockers","None") not in ("None", ""):
            actions.append({
                "action_id": f"ACT-{action_num:03d}",
                "project": pid,
                "action": f"Resolve: {p.get('blockers','')}",
                "owner": p.get("owner",""),
                "due_date": p.get("end_date",""),
                "priority": "HIGH" if p.get("rag_status") == "RED" else "MEDIUM",
                "status": "OPEN",
            })
            action_num += 1

        # Add pending milestones
        for ms in p.get("milestones",[]):
            if not ms.get("done", False):
                actions.append({
                    "action_id": f"ACT-{action_num:03d}",
                    "project": pid,
                    "action": ms.get("name",""),
                    "owner": p.get("owner",""),
                    "due_date": ms.get("due",""),
                    "priority": "HIGH" if p.get("rag_status") == "RED" else "MEDIUM",
                    "status": "OPEN",
                })
                action_num += 1

    for ri, act in enumerate(actions, 2):
        row_data = [act["action_id"], act["project"], act["action"],
                    act["owner"], act["due_date"], act["priority"], act["status"]]
        for ci, v in enumerate(row_data, 1):
            cell = ws3.cell(row=ri, column=ci, value=v)
            cell.font = bfont(); cell.alignment = left(); cell.border = bdr
            if ci == 6:   # priority
                cell.font = Font(name="Calibri", bold=True,
                                 color=C_RED if v == "HIGH" else C_AMBER, size=10)
        set_data(ws3, ri, len(act_cols), alt=(ri%2==0))
        ws3.row_dimensions[ri].height = 16

    ws3.auto_filter.ref = f"A1:{get_column_letter(len(act_cols))}1"
    print(f"  ✓ Sheet 3: Action Log ({len(actions)} actions)")

    # Save
    wb.save(REPORT_PATH)
    size = REPORT_PATH.stat().st_size / 1024
    print(f"\n✅ PMO Tracker saved: {REPORT_PATH.name} ({size:.0f} KB)")
    print(f"   Sheets: {' | '.join(wb.sheetnames)}\n")


if __name__ == "__main__":
    generate()
